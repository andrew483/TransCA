# 广义网络，有共乘出行
import copy
import datetime
import math
import time
import numpy as np
import pandas as pd
from openpyxl import Workbook


theta = 1  # Logit公式的参数


class TransNet:
    # 交通网络的类，直接以向量存储link和path

    ######定义交通网络类中的基本属性######
    Link_Num = 0  # 路段数量，76条路段
    Capacity = 0  # 路段容量
    Length = 0  # 路段长度
    FFT = 0  # 路段自由流时间
    Route_Num = 0  # 路径数量（不区分角色），528个OD对，528×1，每一项都=10
    Role_Num = 0  # 交通网络中的角色数量，7个角色
    Path_Link = 0  # 路径和路段的转换矩阵，（5280×7）×76
    Link_Path = 0  # 路段和路径的转换矩阵，76×（5280×7）
    Car_Path_Link = 0  # 小汽车的 路径-路段转换矩阵
    Drive_Path_Link = 0  # 驾车模式的 路径-路段转换矩阵
    PT_Path_Link = 0  # 公交模式的 路径-路段转换矩阵
    OD_Num = 0  # OD对数量，528
    OD_Demand_C = 0  # 交通网络中的OD需求量（有车的人）
    OD_Demand_NC = 0  # 交通网络中的OD需求量（没车的人）
    parameters_network = []  # 交通网络中的数据设定

    # VOT [0:7] COI [7:11] 共乘时间费 [11:15] 共乘里程费 [15:19] 浮动定价参数 [19:23] 固定费 [23] 使用费 [24] 公交里程费 [25]

    ######定义交通网络类中的类函数######
    def InitialNet(self, filename, rolenum, parameters_network):  # 类的初始化操作，类似于构造函数，构造出一个空网络
        self.Role_Num = rolenum  # 获取角色的数量
        self.parameters_network = parameters_network  # 获取网络参数信息

        # 在 ”Link“ 中获取路段的自由流时间、容量、长度，形成列向量
        FFT = pd.read_excel(filename, sheet_name="Link", usecols=[3], skiprows=0)  # 路段的自由流时间
        self.FFT = FFT.values
        Cap = pd.read_excel(filename, sheet_name="Link", usecols=[4], skiprows=0)  # 路段的容量
        self.Capacity = Cap.values
        Length = pd.read_excel(filename, sheet_name="Link", usecols=[5], skiprows=0)  # 路段的长度
        self.Length = Length.values

        # 在 "OD" 中获取网络的OD对信息
        demand_C = pd.read_excel(filename, sheet_name="OD", usecols=[4], skiprows=0)  # 有车用户的需求
        self.OD_Demand_C = demand_C.values
        demand_NC = pd.read_excel(filename, sheet_name="OD", usecols=[5], skiprows=0)  # 无车用户的需求
        self.OD_Demand_NC = demand_NC.values
        self.OD_Num = self.OD_Demand_C.shape[0]  # OD对数量 = 528

        # 获取 路径-路段矩阵，“Route_Link”
        Route_Link = pd.read_excel(filename, sheet_name="Route-Link", header=None, usecols=range(76),
                                   skiprows=0)  # （不区分角色）路径-路段转换矩阵，5280×76
        Route_Link = Route_Link.values
        self.Link_Num = Route_Link.shape[1]  # 路段数量 = 76
        Route_Num = Route_Link.shape[0]  # 路径数量 = 5280
        self.Route_Num = 10 * np.ones([self.OD_Num, 1], dtype=int)  # 每个OD对的路径数量 = 10
        self.Path_Link = np.zeros([Route_Num * rolenum, self.Link_Num], dtype=int)  # （包含角色）路径-路段转换矩阵，（5280×7）×76
        for i in range(Route_Num):
            for k in range(rolenum):
                self.Path_Link[rolenum * i + k,] = Route_Link[i,]
        self.Path_Link = np.mat(self.Path_Link)  # （包含角色）路径-路段转换矩阵，（5280×7）×76
        self.Link_Path = self.Path_Link.T  # （包含角色）路径-路段转换矩阵，76×（5280×7）

        # 计算 Car、Drive、PT 不同模式特定的 路径-路段转换矩阵
        temp = np.mat(np.zeros([1, self.Link_Num]))
        self.Car_Path_Link = copy.deepcopy(self.Path_Link)  # Car：SD，RD
        self.Drive_Path_Link = copy.deepcopy(self.Path_Link)  # Drive: SD，RD，R
        self.PT_Path_Link = copy.deepcopy(self.Path_Link)  # PT
        for i in range(self.OD_Num):
            for j in range(10):
                for kc in range(3, 7):
                    self.Car_Path_Link[70 * i + self.Role_Num * j + kc,] = temp
                for kd in range(5, 7):
                    self.Drive_Path_Link[70 * i + self.Role_Num * j + kd,] = temp
                for kp in range(5):
                    self.PT_Path_Link[70 * i + self.Role_Num * j + kp,] = temp

    def Initial_PathFlow(self):
        pathflow_0 = np.mat(np.zeros([self.OD_Num * 70, 1]))
        od_pathflow_0 = np.mat(np.zeros([70, 1]))
        for od_i in range(self.OD_Num):
            for i in range(10):
                # 有车用户
                od_pathflow_0[0 + self.Role_Num * i, 0] = self.OD_Demand_C[od_i, 0] / 50  # SD
                od_pathflow_0[1 + self.Role_Num * i, 0] = self.OD_Demand_C[od_i, 0] / 50  # RD_C
                od_pathflow_0[2 + self.Role_Num * i, 0] = self.OD_Demand_C[od_i, 0] / 50  # RD_NC
                od_pathflow_0[3 + self.Role_Num * i, 0] = self.OD_Demand_C[od_i, 0] / 50  # R_C
                od_pathflow_0[5 + self.Role_Num * i, 0] = self.OD_Demand_C[od_i, 0] / 50  # PT_C
                # 无车用户
                od_pathflow_0[4 + self.Role_Num * i, 0] = od_pathflow_0[2 + self.Role_Num * i, 0]  # R_NC = RD_NC
                od_pathflow_0[6 + self.Role_Num * i, 0] = self.OD_Demand_NC[od_i, 0] / 10 - od_pathflow_0[
                    4 + self.Role_Num * i, 0]  # PT_NC = NC - R_NC
            pathflow_0[70 * od_i: 70 * (od_i + 1), ] = od_pathflow_0
        return pathflow_0

    ## 路径流量是 （528×10×7）×1 的列向量，角色排列分别是  SD，RD_C，RD_NC，R_C，R_NC，PT_C，PT_NC
    def LinkTime(self, pathflow):  # 根据传入的路径流量，计算出某一小汽车和公交乘客的路段流量，并用BPR函数得到路段出行时间
        # Driving mode 的路段出行时间
        Car_Link_Flow = self.Car_Path_Link.T * pathflow  # [76×(5280×7)]×[(5280×7)×1]
        Drive_Link_Time = np.mat(np.zeros([self.Link_Num, 1]))
        for i in range(self.Link_Num):
            Drive_Link_Time[i, 0] = self.FFT[i, 0] * (1 + 0.15 * (Car_Link_Flow[i, 0] / self.Capacity[i, 0]) ** 4)

        # PT mode 的路段出行时间
        PT_Link_Flow = self.PT_Path_Link.T * pathflow  # [76×(5280×7)]×[(5280×7)×1]
        PT_Link_Time = np.mat(np.zeros([self.Link_Num, 1]))
        for i in range(self.Link_Num):  # 应用BPR函数计算公交车行驶时间，假设平均每辆公交车上8个乘客，1辆公交车 = 1.5辆小汽车
            PT_Link_Time[i, 0] = self.FFT[i, 0] * (
                    1 + 0.15 * ((PT_Link_Flow[i, 0] / 8) / (self.Capacity[i, 0] / 1.5)) ** 4)

        LinkTime = [Drive_Link_Time, PT_Link_Time]  # 路段出行时间列表
        return LinkTime

    def ODFlow(self, pathflow):  # 根据传入的流量，统计OD对间不同角色的流量
        od_flow = np.mat(np.zeros([self.OD_Num * self.Role_Num, 1]))  # (528×7)×1，7个角色
        # pathflow，（528×10×7）×1
        for od_i in range(self.OD_Num):  # 528
            for i in range(self.Role_Num):  # 7
                for j in range(10):
                    od_flow[7 * od_i + i, 0] += pathflow[70 * od_i + i + j * self.Role_Num, 0]
        return od_flow

    def PathTime(self, LinkTime):  # 计算OD对间路径的 出行时间
        path_time = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        Drive_path_time = self.Drive_Path_Link * LinkTime[0]  # Driving mode 的路径出行时间，[(528×10×7)×76]×[76×1]
        PT_path_time = self.PT_Path_Link * LinkTime[1]  # PT mode 的路径出行时间，[(528×10×7)×76]×[76×1]
        path_time = Drive_path_time + PT_path_time  # 各出行方式的 路径出行时间，(528×70)×1
        return path_time

    def TimeCost(self, path_time):  # 根据不同角色的VOT（1×7），计算某一OD对路径的 TimeCost
        time_cost = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        VOT = np.mat(np.array([[self.parameters_network[0]], [self.parameters_network[1]], [self.parameters_network[2]], \
                               [self.parameters_network[3]], [self.parameters_network[4]], [self.parameters_network[5]], \
                               [self.parameters_network[6]]]))  # 不同角色的VOT
        for od_i in range(self.OD_Num):
            for i in range(self.Role_Num):
                for j in range(10):
                    time_cost[70 * od_i + i + j * self.Role_Num, 0] = VOT[i, 0] * path_time[
                        70 * od_i + i + j * self.Role_Num, 0]  # 时间成本 = 时间 × VOT
        return time_cost

    def IncvCost(self, path_time):  # 计算某一OD对路径的 Inconvenience Cost
        incv_cost = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):
            for i in range(10):
                incv_cost[70 * od_i + 1 + i * self.Role_Num, 0] = self.parameters_network[7] * path_time[
                    70 * od_i + 1 + i * self.Role_Num, 0]
                incv_cost[70 * od_i + 2 + i * self.Role_Num, 0] = self.parameters_network[8] * path_time[
                    70 * od_i + 2 + i * self.Role_Num, 0]
                incv_cost[70 * od_i + 3 + i * self.Role_Num, 0] = self.parameters_network[9] * path_time[
                    70 * od_i + 3 + i * self.Role_Num, 0]
                incv_cost[70 * od_i + 4 + i * self.Role_Num, 0] = self.parameters_network[10] * path_time[
                    70 * od_i + 4 + i * self.Role_Num, 0]
        return incv_cost

    # 时长费0.4，里程费0.2，浮动定价系数设为6和2
    def Price(self, path_time, od_flow):  # 计算共乘收费中的收费
        path_length = self.Path_Link * self.Length  # [(5280×70)×76]×[76×1]

        TimeFee = path_time
        MilageFee = path_length
        SurgeFee = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))

        for od_i in range(self.OD_Num):
            for i in range(10):
                ##里程费
                MilageFee[70 * od_i + 0 + i * self.Role_Num, 0] = 0
                MilageFee[70 * od_i + 1 + i * self.Role_Num, 0] *= -1 * self.parameters_network[15]
                MilageFee[70 * od_i + 2 + i * self.Role_Num, 0] *= -1 * self.parameters_network[16]
                MilageFee[70 * od_i + 3 + i * self.Role_Num, 0] *= self.parameters_network[17]
                MilageFee[70 * od_i + 4 + i * self.Role_Num, 0] *= self.parameters_network[18]
                MilageFee[70 * od_i + 5 + i * self.Role_Num, 0] = 0
                MilageFee[70 * od_i + 6 + i * self.Role_Num, 0] = 0

                ##时间费
                TimeFee[70 * od_i + 0 + i * self.Role_Num, 0] = 0
                TimeFee[70 * od_i + 1 + i * self.Role_Num, 0] *= -1 * self.parameters_network[11]
                TimeFee[70 * od_i + 2 + i * self.Role_Num, 0] *= -1 * self.parameters_network[12]
                TimeFee[70 * od_i + 3 + i * self.Role_Num, 0] *= self.parameters_network[13]
                TimeFee[70 * od_i + 4 + i * self.Role_Num, 0] *= self.parameters_network[14]
                TimeFee[70 * od_i + 5 + i * self.Role_Num, 0] = 0
                TimeFee[70 * od_i + 6 + i * self.Role_Num, 0] = 0

                ##浮动定价
                SurgeFee[70 * od_i + 1 + i * self.Role_Num, 0] = self.parameters_network[19] * od_flow[7 * od_i + 1, 0]
                SurgeFee[70 * od_i + 2 + i * self.Role_Num, 0] = self.parameters_network[20] * od_flow[7 * od_i + 2, 0]
                SurgeFee[70 * od_i + 3 + i * self.Role_Num, 0] = self.parameters_network[21] * od_flow[7 * od_i + 3, 0]
                SurgeFee[70 * od_i + 4 + i * self.Role_Num, 0] = self.parameters_network[22] * od_flow[7 * od_i + 4, 0]

        ##总的费用 = 里程费+时间费+浮动定价
        Price = MilageFee + TimeFee + SurgeFee
        return Price

    def TotalCost(self, path_time, od_flow):  # 计算总的出行费用
        # 车主要承担车辆保养的费用，假设为1；司机承担车辆使用的损耗，假设为1；公交车乘客要花票价，假设为 0.4*里程长度
        other_cost = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        PT_path_length = self.PT_Path_Link * self.Length  # [(5280×70)×76]×[76×1]
        PT_MileFee = self.parameters_network[25] * PT_path_length
        for od_i in range(self.OD_Num):
            for i in range(10):
                other_cost[70 * od_i + 0 + i * self.Role_Num, 0] = self.parameters_network[23] + \
                                                                   self.parameters_network[24]
                other_cost[70 * od_i + 1 + i * self.Role_Num, 0] = self.parameters_network[23] + \
                                                                   self.parameters_network[24]
                other_cost[70 * od_i + 2 + i * self.Role_Num, 0] = self.parameters_network[23] + \
                                                                   self.parameters_network[24]
                other_cost[70 * od_i + 3 + i * self.Role_Num, 0] = self.parameters_network[23]
                other_cost[70 * od_i + 5 + i * self.Role_Num, 0] = PT_MileFee[70 * od_i + 5 + i * self.Role_Num, 0] + \
                                                                   self.parameters_network[23]
                other_cost[70 * od_i + 6 + i * self.Role_Num, 0] = PT_MileFee[70 * od_i + 6 + i * self.Role_Num, 0]

        # 将各项相加费用相加起来
        total_cost = self.TimeCost(path_time) + self.IncvCost(path_time) + self.Price(path_time, od_flow) + other_cost
        return total_cost

    def Multiplier(self, total_cost):  # 计算网络中各个OD对、各个路径的乘子 λ2，λ3
        Lambda = np.mat(np.zeros([10 * self.OD_Num, 2]))  # 乘子 λ，(10×528)×2
        # 先计算 λ(p, 2) = 0.5 * [C(p, 4) - C(p, 2)]
        for od_i in range(self.OD_Num):  # total_cost，(10×7×528)×1
            for i in range(10):
                Lambda[10 * od_i + i, 0] = (1 / 2) * (total_cost[70 * od_i + 3 + i * self.Role_Num, 0] - total_cost[
                    70 * od_i + 1 + i * self.Role_Num, 0])
        # 再计算 λ(p, 3)
        t_1 = np.mat(np.zeros([self.OD_Num, 1]))  # 528×1
        t_2 = np.mat(np.zeros([self.OD_Num, 1]))
        t_3 = np.mat(np.zeros([self.OD_Num, 1]))
        for od_i in range(self.OD_Num):  # total_cost，(10×7×528)×1
            for i in range(10):
                t_1[od_i, 0] += math.exp(-theta * total_cost[70 * od_i + 0 + i * self.Role_Num, 0])
                t_1[od_i, 0] += 2 * math.exp((-theta / 2) * (
                        total_cost[70 * od_i + 1 + i * self.Role_Num, 0] + total_cost[
                    70 * od_i + 3 + i * self.Role_Num, 0]))
                t_1[od_i, 0] += math.exp(-theta * total_cost[70 * od_i + 5 + i * self.Role_Num, 0])
                t_2[od_i, 0] += math.exp(-theta * total_cost[70 * od_i + 6 + i * self.Role_Num, 0])
                t_3[od_i, 0] += math.exp((-theta / 2) * (total_cost[70 * od_i + 2, 0] + total_cost[70 * od_i + 4, 0]
                                                         + total_cost[70 * od_i + 2 + i * self.Role_Num, 0] +
                                                         total_cost[70 * od_i + 4 + i * self.Role_Num, 0]))
        # A (x*x) + B (x) + C = 0 , 一元二次方程
        A = np.mat(np.zeros([self.OD_Num, 1]))  # 528×1
        B = np.mat(np.zeros([self.OD_Num, 1]))
        C = np.mat(np.zeros([self.OD_Num, 1]))
        delta = np.mat(np.zeros([self.OD_Num, 1]))  # △ = B*B - 4*A*C
        for od_i in range(self.OD_Num):
            A[od_i, 0] = self.OD_Demand_NC[od_i, 0] * math.exp(-theta * total_cost[70 * od_i + 4, 0]) * t_1[od_i, 0]
            C[od_i, 0] = - self.OD_Demand_C[od_i, 0] * math.exp(-theta * total_cost[70 * od_i + 2, 0]) * t_2[od_i, 0]
            B[od_i, 0] = (self.OD_Demand_NC[od_i, 0] - self.OD_Demand_C[od_i, 0]) * t_3[od_i, 0]
            delta[od_i, 0] = B[od_i, 0] ** 2 - 4 * A[od_i, 0] * C[od_i, 0]
            temp = (-B[od_i, 0] + (delta[od_i, 0] ** 0.5)) / (2 * A[od_i, 0])  # 求根公式得到的正解，＞0
            Lambda[10 * od_i + 0, 1] = math.log(temp) / theta
        # λ(pi, 3) = λ(p1, 3) - 0.5 * [C(p1, 5) - C(p1, 3) - C(pi, 5) + C(pi, 3)]
        for od_i in range(self.OD_Num):
            for i in range(1, 10):
                Lambda[10 * od_i + i, 1] = Lambda[10 * od_i + 0, 1] - (1 / 2) * (
                    (total_cost[70 * od_i + 4, 0] - total_cost[70 * od_i + 2, 0])) \
                                           + (1 / 2) * (total_cost[70 * od_i + 4 + i * self.Role_Num] - total_cost[
                    70 * od_i + 2 + i * self.Role_Num])

        return Lambda

    def GeneralCost(self, total_cost, Lambda):  # 计算 OD 对之间各个路径的广义出行费用
        #  计算网络中各个角色的广义出行费用
        theta = 1  # Logit 模型的系数 θ=1
        general_cost = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):
            for i in range(10):
                general_cost[70 * od_i + 0 + i * self.Role_Num, 0] = total_cost[70 * od_i + 0 + i * self.Role_Num, 0]

                general_cost[70 * od_i + 1 + i * self.Role_Num, 0] = total_cost[70 * od_i + 1 + i * self.Role_Num, 0] + \
                                                                     Lambda[10 * od_i + i, 0]

                general_cost[70 * od_i + 2 + i * self.Role_Num, 0] = total_cost[70 * od_i + 2 + i * self.Role_Num, 0] + \
                                                                     Lambda[10 * od_i + i, 1]

                general_cost[70 * od_i + 3 + i * self.Role_Num, 0] = total_cost[70 * od_i + 3 + i * self.Role_Num, 0] - \
                                                                     Lambda[10 * od_i + i, 0]

                general_cost[70 * od_i + 4 + i * self.Role_Num, 0] = total_cost[70 * od_i + 4 + i * self.Role_Num, 0] - \
                                                                     Lambda[10 * od_i + i, 1]

                general_cost[70 * od_i + 5 + i * self.Role_Num, 0] = total_cost[70 * od_i + 5 + i * self.Role_Num, 0]

                general_cost[70 * od_i + 6 + i * self.Role_Num, 0] = total_cost[70 * od_i + 6 + i * self.Role_Num, 0]

        return general_cost

    def AssignFlow(self, general_cost):  # 根据广义出行成本，分配各个 OD 对之间的路径流量
        pathflow = np.mat(np.zeros([self.OD_Num * 10 * self.Role_Num, 1]))
        theta = 1  # θ=1
        for od_i in range(self.OD_Num):
            od_demand_C = self.OD_Demand_C[od_i, 0]  # 交通网络中的OD需求量（有车的人）
            od_demand_NC = self.OD_Demand_NC[od_i, 0]  # 交通网络中的OD需求量（没车的人）
            sum_C = 0  # Logit分配模型的分母
            sum_NC = 0
            for i in range(10):  # 计算分母
                #  有车用户
                sum_C += math.exp(-theta * general_cost[70 * od_i + 0 + i * self.Role_Num, 0])
                sum_C += math.exp(-theta * general_cost[70 * od_i + 1 + i * self.Role_Num, 0])
                sum_C += math.exp(-theta * general_cost[70 * od_i + 2 + i * self.Role_Num, 0])
                sum_C += math.exp(-theta * general_cost[70 * od_i + 3 + i * self.Role_Num, 0])
                sum_C += math.exp(-theta * general_cost[70 * od_i + 5 + i * self.Role_Num, 0])
                # 无车用户
                sum_NC += math.exp(-theta * general_cost[70 * od_i + 4 + i * self.Role_Num, 0])
                sum_NC += math.exp(-theta * general_cost[70 * od_i + 6 + i * self.Role_Num, 0])

            for i in range(10):  # 依据 Logit 模型 分配流量【一开始需求没乘，我去(⇀‸↼‶)】
                # 有车用户
                pathflow[70 * od_i + 0 + i * self.Role_Num, 0] = od_demand_C * math.exp(
                    -theta * general_cost[70 * od_i + 0 + i * self.Role_Num, 0]) / sum_C
                pathflow[70 * od_i + 1 + i * self.Role_Num, 0] = od_demand_C * math.exp(
                    -theta * general_cost[70 * od_i + 1 + i * self.Role_Num, 0]) / sum_C
                pathflow[70 * od_i + 2 + i * self.Role_Num, 0] = od_demand_C * math.exp(
                    -theta * general_cost[70 * od_i + 2 + i * self.Role_Num, 0]) / sum_C
                pathflow[70 * od_i + 3 + i * self.Role_Num, 0] = od_demand_C * math.exp(
                    -theta * general_cost[70 * od_i + 3 + i * self.Role_Num, 0]) / sum_C
                pathflow[70 * od_i + 5 + i * self.Role_Num, 0] = od_demand_C * math.exp(
                    -theta * general_cost[70 * od_i + 5 + i * self.Role_Num, 0]) / sum_C
                # 无车用户
                pathflow[70 * od_i + 4 + i * self.Role_Num, 0] = od_demand_NC * math.exp(
                    -theta * general_cost[70 * od_i + 4 + i * self.Role_Num, 0]) / sum_NC
                pathflow[70 * od_i + 6 + i * self.Role_Num, 0] = od_demand_NC * math.exp(
                    -theta * general_cost[70 * od_i + 6 + i * self.Role_Num, 0]) / sum_NC

        return pathflow

    def CostAve(self, pathflow_0):  # CA算法，迭代不动点问题
        elta = 0.01  # 迭代精度
        #  由初始流量，计算初始的广义出行成本
        LinkTime_0 = self.LinkTime(pathflow_0)
        path_time_0 = self.PathTime(LinkTime_0)
        od_flow_0 = self.ODFlow(pathflow_0)
        total_cost_0 = self.TotalCost(path_time_0, od_flow_0)
        lambda_0 = self.Multiplier(total_cost_0)
        general_cost_0 = self.GeneralCost(total_cost_0, lambda_0)

        general_cost_old = copy.deepcopy(general_cost_0)
        iter = 0
        while iter < 10 ** 6:
            # 迭代过程
            iter += 1
            stepsize = 1 / iter  # 步长表达式
            pathflow_mid = self.AssignFlow(general_cost_old)  # 路径流量（这个应该是符合流量关系的？）
            LinkTime_mid = self.LinkTime(pathflow_mid)  # 路段时间
            path_time_mid = self.PathTime(LinkTime_mid)  # 路径时间
            od_flow_mid = self.ODFlow(pathflow_mid)  # OD的角色流量
            total_cost_mid = self.TotalCost(path_time_mid, od_flow_mid)  # 总成本
            lambda_mid = self.Multiplier(total_cost_mid)  # 乘子
            general_cost_mid = self.GeneralCost(total_cost_mid, lambda_mid)  # 广义出行成本
            general_cost_new = general_cost_old + stepsize * (general_cost_mid - general_cost_old)  # 不动点问题计算公式

            # 检查收敛条件
            temp_old = 0
            temp_dif = 0
            temp_new = 0
            temp_mid = 0
            for i in range(self.OD_Num * self.Role_Num * 10):
                temp_old += general_cost_old[i, 0] ** 2
                temp_new += general_cost_new[i, 0] ** 2
                temp_mid += general_cost_mid[i, 0] ** 2
                temp_dif += (general_cost_new[i, 0] - general_cost_old[i, 0]) ** 2
            # print("old = ", temp_old, "  new = ", temp_new, "  mid = ", temp_mid, "  dif = ", temp_dif)
            accuracy = temp_dif ** 0.5  # 绝对的收敛条件
            if accuracy <= elta:  # 满足收敛条件，停止迭代 10 * (-4)
                # print(accuracy)
                # print("old = ", temp_old, "  new = ", temp_new, "  mid = ", temp_mid, "  dif = ", temp_dif)
                break
            else:
                general_cost_old = copy.deepcopy(general_cost_new)  # 更新广义出行成本，继续迭代

        pathflow_new = self.AssignFlow(general_cost_new)  # 由结果广义出行成本，生成结果路径流量
        # print("迭代次数 = ", iter)
        return [pathflow_new, general_cost_new]

    def CheckResult(self, pathflow, general_cost):  # 检查最终结果
        Pi = np.mat(np.zeros([self.OD_Num * 10, 7]))  # (528×10) × 7
        for od_i in range(self.OD_Num):  # 528
            for i in range(10):
                for j in range(self.Role_Num):  # 7
                    Pi[10 * od_i + i, j] = general_cost[70 * od_i + i * self.Role_Num + j, 0] \
                                           + (1 / theta) * math.log(pathflow[70 * od_i + i * self.Role_Num + j, 0])
        return Pi


if __name__ == '__main__':

    filename1 = r"../data/SiouxFalls_Network.xlsx"
    filename2 = r"../results/save.xlsx"
    # VOT [0:7] COI [7:11] 共乘时间费 [11:15] 共乘里程费 [15:19] 浮动定价参数 [19:23] 固定费 [23] 使用费 [24] 公交里程费 [25]
    parameters_network = [1, 0.8, 0.8, 0.4, 0.3, 0.05, 0.03, 0.3, 0.3, 0.2, 0.2, 0.4, 0.4, 0.4, 0.4, 0.2, 0.2,
                          0.2, 0.2, 0.5, 0.5, 0.3, 0.3, 3, 5, 0.4]
    SiouxFall = TransNet()
    SiouxFall.InitialNet(filename=filename1, rolenum=7, parameters_network=parameters_network)
    time_1 = time.time()

    pathflow_0 = SiouxFall.Initial_PathFlow()  # 初始路径流量
    [pathflow, general_cost] = SiouxFall.CostAve(pathflow_0)  # 结果路径流量，结果路径出行总成本
    Pi = SiouxFall.CheckResult(pathflow, general_cost)  # Pi 乘子的结果
    wb = Workbook()

    # 将流量结果写入文件中
    # SD:单独开车	RD_C:接到‘有车的乘客’的网约车司机 RD_NC:接到’没车的乘客‘的网约车司机	R_C:有车的网约车乘客
    # R_NC:没车的网约车乘客 PT_C:有车的公交乘客 PT_NC:没车的公交乘客
    ws1 = wb.create_sheet("Path Flow Result")
    for od_i in range(SiouxFall.OD_Num):  # 总共 528*10 行
        ws1.append(["SD", "RD_C", "RD_NC", "R_C", "R_NC", "PT_C", "PT_NC"])
        for i in range(10):
            temp_od_pathflow = []
            for j in range(SiouxFall.Role_Num):
                temp_od_pathflow.append(pathflow[70 * od_i + i * SiouxFall.Role_Num + j, 0])
            ws1.append(temp_od_pathflow)

    # 将 π 结果写入文件中
    ws2 = wb.create_sheet("Path Flow 2-Norm")
    for od_i in range(SiouxFall.OD_Num):  # 总共 528*10 行
        ws2.append(["SD", "RD_C", "RD_NC", "R_C", "R_NC", "PT_C", "PT_NC"])
        for i in range(10):
            temp_Pi = []
            for j in range(SiouxFall.Role_Num):
                temp_Pi.append(Pi[10 * od_i + i, j])
            ws2.append(temp_Pi)
    wb.save(filename2)

    # 输出运行成功提示
    time_2 = time.time()
    message = "Save Path：{} \n" \
              "Calculate Time：{} s\n" \
              "Start Time：{} →  End Time：{}" \
        .format(filename2, "%.2f" % (time_2 - time_1), time.strftime('%H:%M:%S', time.localtime(time_1)),
                time.strftime('%H:%M:%S', time.localtime(time_2)))
    print(message)            
